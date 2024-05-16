import os
from flask import Flask, render_template, request, redirect, url_for
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

class Task:
    def __init__(self, title, description):
        self.title = title
        self.description = description[:80] + '...' if len(description) > 80 else description

tasks = []

def load_tasks():
    global tasks
    if not tasks:
        if os.path.exists("tasks.xlsx"):
            workbook = load_workbook("tasks.xlsx")
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                title, description = row
                tasks.append(Task(title, description))
            workbook.close()

def save_tasks():
    workbook = Workbook()
    sheet = workbook.active
    for i, task in enumerate(tasks, start=1):
        sheet.cell(row=i, column=1, value=task.title)
        sheet.cell(row=i, column=2, value=task.description)
    workbook.save("tasks.xlsx")
    workbook.close()

@app.route('/', methods=['GET', 'POST'])
def index():
    global tasks
    if request.method == 'POST':
        title = request.form['task_title']
        description = request.form['task_description']
        tasks.append(Task(title, description))
        save_tasks()
    else:
        load_tasks()
    return render_template('index.html', tasks=list(enumerate(tasks)))

@app.route('/remove/<int:task_id>', methods=['POST'])
def remove_task(task_id):
    global tasks
    del tasks[task_id]
    save_tasks()
    return redirect(url_for('index'))

if __name__ == "__main__":
    app.run(debug=True)